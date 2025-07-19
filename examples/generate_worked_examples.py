"""
Example 1: The Responsible Weekend Getaway (Single Currency)
Participants: Alice, Bob, Charlie, Diana

Test Transactions:
Date        | Description                                    | Amount | Paid By | Participants
2024-03-15  | Airbnb                                         | 240.00 | Alice   | All
2024-03-15  | Group dinner                                   | 120.00 | Bob     | All
2024-03-16  | Museum tickets                                 | 60.00  | Charlie | Alice, Charlie, Diana
2024-03-16  | Emergency coffee run                           | 28.00  | Charlie | All
2024-03-16  | Uber                                           | 32.00  | Alice   | Alice, Bob
2024-03-17  | Fancy brunch                                   | 80.00  | Diana   | All
2024-03-17  | Snacks for the road trip home                  | 36.00  | Bob     | All

Expected Results:
Total Expenses: $596.00
Per-person average: $149.00

Expected Final Balances:
Alice: Should receive $110.00
Bob: Should receive $14.00
Charlie: Should pay $58.00
Diana: Should pay $66.00

Expected Optimized Settlement (One potential solution):
Charlie pays Alice: $44.00
Diana pays Alice: $66.00
Charlie pays Bob: $14.00
(Total: 3 transactions instead of 6)
"""

import openpyxl
import os
#import sys
#sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '../')))
from settlesheet import create_expense_spreadsheet, create_optimized_settlement_formulas


#os.makedirs("examples", exist_ok=True)
filename1 = "responsible_weekend_getaway.xlsx"
participants1 = ["Alice", "Bob", "Charlie", "Diana"]
expenses1 = [
    ("2024-03-15", "Airbnb", 240.00, "Alice", "All"),
    ("2024-03-15", "Group dinner", 120.00, "Bob", "All"),
    ("2024-03-16", "Museum tickets", 60.00, "Charlie", "Alice, Charlie, Diana"),
    ("2024-03-16", "Emergency coffee run", 28.00, "Charlie", "All"),
    ("2024-03-16", "Uber", 32.00, "Alice", "Alice, Bob"),
    ("2024-03-17", "Fancy brunch", 80.00, "Diana", "All"),
    ("2024-03-17", "Snacks for the road trip home", 36.00, "Bob", "All"),
]
create_expense_spreadsheet(
    filename=filename1,
    participants=participants1,
    expense_rows=len(expenses1),
    exchange_rates=None,
    color_theme="neutral"
)
wb1 = openpyxl.load_workbook(filename1)
ws1 = wb1.active
for row in ws1.iter_rows(min_row=1, max_row=ws1.max_row):
    if row[0].value and "Date" in str(row[0].value):
        expense_start_row = row[0].row + 1
        break
for i, (date, desc, amount, paid_by, part) in enumerate(expenses1):
    row = expense_start_row + i
    ws1.cell(row=row, column=1).value = date
    ws1.cell(row=row, column=2).value = desc
    ws1.cell(row=row, column=3).value = amount
    ws1.cell(row=row, column=4).value = paid_by
    ws1.cell(row=row, column=5).value = part
wb1.save(filename1)

#########################################################################################

"""
Example 2: The Multi-Currency European Adventure
Participants: Alice, Bob, Charlie
Base Currency: USD
Exchange Rates: EUR = 1.10 USD, GBP = 1.27 USD

Test Transactions:
Date        | Description                | Amount Paid | Paid In | Paid By | Participants
2024-06-10  | London hotel               | 180.00      | GBP     | Alice   | All
2024-06-11  | Fish & chips               | 45.00       | GBP     | Bob     | All
2024-06-12  | Train to Paris             | 150.00      | EUR     | Charlie | All
2024-06-13  | Parisian café breakfast    | 60.00       | EUR     | Alice   | All
2024-06-13  | Museum tickets             | 36.00       | EUR     | Bob     | Alice, Bob
2024-06-14  | Airport sandwich           | 24.00       | USD     | Charlie | All

Expected Results:
Total Expenses (USD): $580.35
Per-person average: $193.45

Expected Final Balances:
Alice: Should receive $94.55
Bob: Should pay $103.30
Charlie: Should receive $8.75

Expected Currency Breakdown:
Alice paid: $294.60 USD (£180 + €60 converted)
Bob paid: $96.75 USD (£45 + €36 converted)
Charlie paid: $189.00 USD (€150 + $24)

Expected Optimized Settlement:
Bob pays Alice: $94.55
Bob pays Charlie: $8.75
(Total: 2 transactions instead of 3)
"""

filename2 = "multicurrency_european_adventure.xlsx"
participants2 = ["Alice", "Bob", "Charlie"]
exchange_rates2 = {"EUR": 1.10, "GBP": 1.27}
expenses2 = [
    ("2024-06-10", "London hotel", 180.00, "GBP", "Alice", "All"),
    ("2024-06-11", "Fish & chips", 45.00, "GBP", "Bob", "All"),
    ("2024-06-12", "Train to Paris", 150.00, "EUR", "Charlie", "All"),
    ("2024-06-13", "Parisian café breakfast", 60.00, "EUR", "Alice", "All"),
    ("2024-06-13", "Museum tickets", 36.00, "EUR", "Bob", "Alice, Bob"),
    ("2024-06-14", "Airport sandwich", 24.00, "USD", "Charlie", "All"),
]
create_expense_spreadsheet(
    filename=filename2,
    participants=participants2,
    expense_rows=len(expenses2),
    exchange_rates=exchange_rates2,
    native_currency="USD",
    color_theme="dark"
)
wb2 = openpyxl.load_workbook(filename2)
ws2 = wb2.active
for row in ws2.iter_rows(min_row=1, max_row=ws2.max_row):
    if row[0].value and "Date" in str(row[0].value):
        expense_start_row = row[0].row + 1
        break
for i, (date, desc, amount, paid_in, paid_by, part) in enumerate(expenses2):
    row = expense_start_row + i
    ws2.cell(row=row, column=1).value = date
    ws2.cell(row=row, column=2).value = desc
    ws2.cell(row=row, column=3).value = amount
    ws2.cell(row=row, column=4).value = paid_in
    ws2.cell(row=row, column=6).value = paid_by
    ws2.cell(row=row, column=7).value = part
wb2.save(filename2) 

#########################################################################################

"""
Example 3: 8-person Group Reunion Trip
Participants: Alice, Bob, Charlie, Diana, Eve, Frank, Grace, Henry
Demonstration of larger groups

Test Transactions:
Date        | Description           | Amount | Paid By | Participants
2024-11-01  | Hotel rooms (3 nights)| 720.00 | Alice   | All
2024-11-01  | Rental car            | 280.00 | Frank   | Alice, Bob, Charlie, Diana, Eve, Frank
2024-11-02  | Breakfast             | 96.00  | Bob     | All
2024-11-02  | Theme park tickets    | 400.00 | Grace   | Grace, Henry, Alice, Bob, Diana
2024-11-02  | Lunch at park         | 85.00  | Diana   | Grace, Henry, Alice, Bob, Diana
2024-11-03  | Groceries for BBQ     | 120.00 | Charlie | All
2024-11-03  | Wine and beer         | 75.00  | Eve     | Charlie, Diana, Eve, Frank, Grace
2024-11-04  | Gas for rental car    | 48.00  | Frank   | Alice, Bob, Charlie, Diana, Eve, Frank
2024-11-04  | Escape room           | 160.00 | Henry   | Bob, Charlie, Eve, Henry
2024-11-05  | Farewell dinner       | 240.00 | Bob     | All

Expected Results:
Total Paid by Each:
- Alice: $720.00
- Bob: $336.00
- Charlie: $120.00
- Diana: $85.00
- Eve: $75.00
- Frank: $328.00
- Grace: $400.00
- Henry: $160.00
Total: $2,224.00

Final Balances:
- Alice: +$421.04 (is owed)
- Bob: -$2.67 (owes)
- Charlie: -$136.37 (owes)
- Diana: -$228.67 (owes)
- Eve: -$181.66 (owes)
- Frank: +$111.33 (is owed)
- Grace: +$141.00 (is owed)
- Henry: -$124.00 (owes)

Optimized Settlements:
1. Bob pays Alice: $2.67
2. Charlie pays Alice: $136.37
3. Diana pays Alice: $228.67
4. Eve pays Alice: $53.33
5. Eve pays Frank: $111.33
6. Eve pays Grace: $17.00
7. Henry pays Grace: $124.00

Total: 7 transactions
"""

filename3 = "reunion_trip.xlsx"
participants3 = ["Alice", "Bob", "Charlie", "Diana", "Eve", "Frank", "Grace", "Henry"]
expenses3 = [
    ("2024-11-01", "Hotel rooms (3 nights)", 720.00, "Alice", "All"),
    ("2024-11-01", "Rental car", 280.00, "Frank", "Alice, Bob, Charlie, Diana, Eve, Frank"),
    ("2024-11-02", "Breakfast", 96.00, "Bob", "All"),
    ("2024-11-02", "Theme park tickets", 400.00, "Grace", "Grace, Henry, Alice, Bob, Diana"),
    ("2024-11-02", "Lunch at park", 85.00, "Diana", "Grace, Henry, Alice, Bob, Diana"),
    ("2024-11-03", "Groceries for BBQ", 120.00, "Charlie", "All"),
    ("2024-11-03", "Wine and beer", 75.00, "Eve", "Charlie, Diana, Eve, Frank, Grace"),
    ("2024-11-04", "Gas for rental car", 48.00, "Frank", "Alice, Bob, Charlie, Diana, Eve, Frank"),
    ("2024-11-04", "Escape room", 160.00, "Henry", "Bob, Charlie, Eve, Henry"),
    ("2024-11-05", "Farewell dinner", 240.00, "Bob", "All"),
]
create_expense_spreadsheet(
    filename=filename3,
    participants=participants3,
    expense_rows=len(expenses3),
    exchange_rates=None,
    color_theme="sleek"
)
wb3 = openpyxl.load_workbook(filename3)
ws3 = wb3.active
for row in ws3.iter_rows(min_row=1, max_row=ws3.max_row):
    if row[0].value and "Date" in str(row[0].value):
        expense_start_row = row[0].row + 1
        break
for i, (date, desc, amount, paid_by, part) in enumerate(expenses3):
    row = expense_start_row + i
    ws3.cell(row=row, column=1).value = date
    ws3.cell(row=row, column=2).value = desc
    ws3.cell(row=row, column=3).value = amount
    ws3.cell(row=row, column=4).value = paid_by
    ws3.cell(row=row, column=5).value = part
wb3.save(filename3) 
